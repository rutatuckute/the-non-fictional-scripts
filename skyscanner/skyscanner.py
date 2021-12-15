import json
import requests
import pandas as pd
import time
import datetime
import openpyxl
from openpyxl import *
from pandas.io.json import json_normalize

import plotly.express as px
pd.options.plotting.backend = 'plotly'

class QuerySkyscanner:
    
    def __init__(self):
        
        self.generated_at = datetime.datetime.fromtimestamp(time.time()).strftime('%Y%m%d_%H%M%S')
        self.url = "https://skyscanner-skyscanner-flight-search-v1.p.rapidapi.com/apiservices/browsequotes/v1.0/FR/EUR/en-US/PARI-sky/VILN-sky/2020-12-20/2020-12-27"
        self.headers = {
            'x-rapidapi-key': "your_API_key",
            'x-rapidapi-host': "skyscanner-skyscanner-flight-search-v1.p.rapidapi.com"
        }
        
    def request_skyscanner(self):
        
        response = requests.request("GET", self.url, headers=self.headers)
        take_me_home = response.json()
        self.quotes = json_normalize(take_me_home['Quotes'])
        
    def save_to_separate_files(self):
        
        file_path = "C:\\Users\\ruttuc\\Desktop\\take_me_home\\"
        file_save = "/generated_quotes_" + self.generated_at + ".xlsx"
        self.quotes.to_excel(file_path+file_save)
        
    def save_to_unique_file(self):
        
        file_path = "C:\\Users\\ruttuc\\Desktop\\take_me_home\\"
        file_save = "/quotes_hoard.xlsx"
        
        self.quotes['QueryDate'] = self.generated_at
        
        max_row = load_workbook(file_path+file_save)['data'].max_row

        book = load_workbook(file_path+file_save)
        writer = pd.ExcelWriter(file_path+file_save, engine='openpyxl') 
        writer.book = book
        writer.sheets = dict((ws.title, ws) for ws in book.worksheets)

        self.quotes.to_excel(writer, 'data', startrow=max_row, index=False, header=False)
        writer.save()
        
    def generate_visual(self, df, date_column, value_column):
        
        self.fig = df.plot(x=date_column, y=value_column,
                color_discrete_map=dict(zip([value_column], ['orange'])))
        
        self.fig.update_layout(
            width=600,
            height=500,
            plot_bgcolor='black',
            paper_bgcolor='black',
            font=dict(size=13, color='white'),
            margin=dict(l=50, r=20, t=60, b=50),
            legend_title_text='',
            legend = dict(xanchor="left")
        )
        
        self.fig.update_xaxes(title_text='', showgrid=False, linewidth=2)
        self.fig.update_yaxes(title_text='', showgrid=False, linewidth=2)

        self.fig.show()
        



if __name__ == "__main__":
    skyscanner = QuerySkyscanner()
    skyscanner.request_skyscanner()
    skyscanner.save_to_unique_file()

    
